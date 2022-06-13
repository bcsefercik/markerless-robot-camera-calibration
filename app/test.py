import typing
import ipdb

import os
import sys
import json
import pickle
import random
import statistics
from collections import defaultdict

import torch
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import MinkowskiEngine as ME

from utils.transformation import get_base2cam_pose, transform_pose2pose
from utils import config, logger, preprocess, utils, metrics
from utils import calibration as calib_util
from utils.data import get_6_key_points

import data_engine
from inference_engine import InferenceEngine
from dto import TestResultDTO, RawDTO, CalibrationResultDTO


_config = config.Config()
_logger = logger.Logger().get()

_use_cuda = torch.cuda.is_available()
_device = torch.device("cuda" if _use_cuda else "cpu")


class TestApp:
    def __init__(self, data_source=_config.TEST.data_source) -> None:
        self._data_source = data_engine.PickleDataEngine(data_source, cyclic=False)

        self._inference_engine = InferenceEngine()

        self._gt_base_to_cam_pose = np.array(_config.TEST.gt_base_to_cam_pose, dtype=np.float32)

        self.instance_results = None
        self.position_results = None
        self.overall_results = None

        self.predictions: typing.Dict[str, typing.List[TestResultDTO]] = None
        self.calibration: CalibrationResultDTO = None

        self.clear_results()

        random.seed(_config.TEST.seed)
        np.random.seed(_config.TEST.seed)
        torch.manual_seed(_config.TEST.seed)

    def clear_results(self):
        self.instance_results = defaultdict(dict)
        self.position_results = defaultdict(dict)
        self.overall_results = defaultdict(list)

        self.predictions = defaultdict(list)
        self.calibration = None

    def run_tests(self):
        self.clear_results()

        # Make predictions
        with torch.no_grad():
            while True:
                data: RawDTO = self._data_source.get_raw()

                if data is None:
                    break

                data_key = (
                    f"{data.other['position']}/{data.other['filepath'].split('/')[-1]}"
                )
                self.instance_results[data_key]['position'] = data.other['position']

                rgb = preprocess.normalize_colors(data.rgb)  # never use data.rgb below

                seg_results = data.segmentation

                if _config.TEST.SEGMENTATION.evaluate:
                    seg_results = self._inference_engine.predict_segmentation(
                        data.points, rgb
                    )
                    segmentation_metrics = metrics.compute_segmentation_metrics(
                        data.segmentation,
                        seg_results,
                        classes=_config.INFERENCE.SEGMENTATION.classes,
                    )
                    self.instance_results[data_key]['segmentation'] = segmentation_metrics

                result_dto = TestResultDTO(segmentation=seg_results)

                ee_idx = np.where(seg_results == 2)[0]
                ee_raw_points = data.points[ee_idx]  # no origin offset
                ee_raw_rgb = torch.from_numpy(rgb[ee_idx]).to(dtype=torch.float32)
                ee_centered_points, _ = preprocess.center_at_origin(ee_raw_points)

                ee_gt_idx = np.where(data.segmentation == 2)[0]
                ee_gt_raw_points = data.points[ee_gt_idx]  # no origin offset
                ee_gt_centered_points, _ = preprocess.center_at_origin(ee_gt_raw_points)
                ee_centered_points = ee_gt_centered_points

                rot_result = self._inference_engine.predict_rotation(
                    ee_raw_points, ee_raw_rgb
                )

                pos_result, _ = self._inference_engine.predict_translation(
                    ee_raw_points, ee_raw_rgb, q=rot_result
                )

                nn_pose = np.concatenate((pos_result, rot_result))
                result_dto.ee_pose = nn_pose

                nn_pose_metrics = metrics.compute_pose_metrics(data.pose, nn_pose)
                self.instance_results[data_key]['dist_position'] = {'nn': nn_pose_metrics['dist_position']}
                self.instance_results[data_key]['angle_diff'] = {'nn': nn_pose_metrics['angle_diff']}

                self.instance_results[data_key]['ADD_nn'] = metrics.compute_ADD_np(ee_centered_points, data.pose, nn_pose)

                nn_pose_icp = self._inference_engine.match_icp(ee_raw_points, nn_pose)
                nn_pose_icp_metrics = metrics.compute_pose_metrics(data.pose, nn_pose_icp)
                self.instance_results[data_key]['dist_position']['nn_icp'] = nn_pose_icp_metrics['dist_position']
                self.instance_results[data_key]['angle_diff']['nn_icp'] = nn_pose_icp_metrics['angle_diff']

                self.instance_results[data_key]['ADD_nn_icp'] = metrics.compute_ADD_np(ee_centered_points, data.pose, nn_pose_icp)

                result_dto.ee_pose = nn_pose_icp

                result_dto.base_pose = get_base2cam_pose(result_dto.ee_pose, data.ee2base_pose)

                base_pose = np.array(result_dto.base_pose, copy=True)
                if self._inference_engine.camera_link_transformation_pose is not None:
                    base_pose = transform_pose2pose(base_pose, self._inference_engine.camera_link_transformation_pose)
                    result_dto.base_pose_camera_link = base_pose

                base_pose_metrics = metrics.compute_pose_metrics(
                    self._gt_base_to_cam_pose,
                    base_pose
                )
                self.instance_results[data_key]['base2cam'] = {
                    'dist_position': base_pose_metrics['dist_position'],
                    'angle_diff': base_pose_metrics['angle_diff']
                }

                kp_gt_coords, kp_gt_idx = get_6_key_points(ee_raw_points, data.pose, switch_w=False)
                kp_coords, kp_classes, kp_confs = self._inference_engine.predict_key_points(
                    ee_raw_points, ee_raw_rgb,
                )
                mean_kp_error = metrics.compute_kp_error(kp_gt_coords, kp_coords, kp_classes)
                self.instance_results[data_key]['mean_kp_error'] = mean_kp_error

                result_dto.key_points = list(zip(kp_classes, kp_coords))

                if len(kp_classes) > 3:
                    kp_pose = self._inference_engine.predict_pose_from_kp(
                        kp_coords, kp_classes
                    )
                    result_dto.key_points_pose = kp_pose
                    kp_pose_metrics = metrics.compute_pose_metrics(data.pose, kp_pose)

                    self.instance_results[data_key]['dist_position']['kp'] = kp_pose_metrics['dist_position']
                    self.instance_results[data_key]['angle_diff']['kp'] = kp_pose_metrics['angle_diff']

                    self.instance_results[data_key]['ADD_kp'] = metrics.compute_ADD_np(ee_centered_points, data.pose, kp_pose)

                    kp_pose_icp = self._inference_engine.match_icp(ee_raw_points, kp_pose)
                    kp_pose_icp_metrics = metrics.compute_pose_metrics(data.pose, kp_pose_icp)
                    self.instance_results[data_key]['dist_position']['kp_icp'] = kp_pose_icp_metrics['dist_position']
                    self.instance_results[data_key]['angle_diff']['kp_icp'] = kp_pose_icp_metrics['angle_diff']

                    self.instance_results[data_key]['ADD_kp_icp'] = metrics.compute_ADD_np(ee_centered_points, data.pose, kp_pose_icp)

                    result_dto.key_points_pose = kp_pose_icp

                    result_dto.key_points_base_pose = get_base2cam_pose(result_dto.key_points_pose, data.ee2base_pose)

                    key_points_base_pose = np.array(result_dto.key_points_base_pose, copy=True)

                    if self._inference_engine.camera_link_transformation_pose is not None:
                        key_points_base_pose = transform_pose2pose(key_points_base_pose, self._inference_engine.camera_link_transformation_pose)
                        result_dto.key_points_base_pose_camera_link = key_points_base_pose

                    key_points_base_pose_metrics = metrics.compute_pose_metrics(
                        self._gt_base_to_cam_pose,
                        key_points_base_pose
                    )
                    self.instance_results[data_key]['base2cam']['dist_position_kp'] = key_points_base_pose_metrics['dist_position']
                    self.instance_results[data_key]['base2cam']['angle_diff_kp'] = key_points_base_pose_metrics['angle_diff']

                result_dto.is_confident = self._inference_engine.check_sanity(
                    data.to_point_cloud_dto(),
                    result_dto,
                    kp_error_margin=_config.TEST.KEY_POINTS.error_margin
                )
                if _config.TEST.ignore_unconfident and not result_dto.is_confident:
                    self.instance_results.pop(data_key)

                result_dto.id = data_key

                self.predictions[data.other['position']].append(result_dto)

                # ipdb.set_trace()
                _logger.info(f'{data_key}{"" if result_dto.is_confident else ", ignored"}')

        # with open('predictions.pickle', 'wb') as fp:
        #     pickle.dump(self.predictions, fp, protocol=pickle.HIGHEST_PROTOCOL)

        # with open('predictions.pickle', 'rb') as fp:
        #     self.predictions = pickle.load(fp)

        self.calibration = self._inference_engine.calibrate(self.predictions)

        # Print resultws to a spreadsheet.
        position_results_raw = defaultdict(list)
        for ir in self.instance_results.values():
            position_results_raw[ir['position']].append(ir)

        for pos, irs in position_results_raw.items():

            self.position_results[pos]['base2cam_dist_position'] = [ir['base2cam']['dist_position'] for ir in irs]
            self.position_results[pos]['base2cam_angle_diff'] = [ir['base2cam']['angle_diff'] for ir in irs]
            self.position_results[pos]['base2cam_dist_position_kp'] = [ir['base2cam']['dist_position_kp'] for ir in irs if 'dist_position_kp' in ir['base2cam']]
            self.position_results[pos]['base2cam_angle_diff_kp'] = [ir['base2cam']['angle_diff_kp'] for ir in irs if 'angle_diff_kp' in ir['base2cam']]

            self.position_results[pos]['mean_kp_error'] = [ir['mean_kp_error'] for ir in irs]

            self.position_results[pos]['angle_diff_nn'] = [ir['angle_diff']['nn'] for ir in irs]
            self.position_results[pos]['angle_diff_nn_icp'] = [ir['angle_diff']['nn_icp'] for ir in irs]
            self.position_results[pos]['angle_diff_kp'] = [ir['angle_diff']['kp'] for ir in irs if 'kp' in ir['dist_position']]
            self.position_results[pos]['angle_diff_kp_icp'] = [ir['angle_diff']['kp_icp'] for ir in irs if 'kp' in ir['dist_position']]

            self.position_results[pos]['dist_position_nn'] = [ir['dist_position']['nn'] for ir in irs]
            self.position_results[pos]['dist_position_nn_icp'] = [ir['dist_position']['nn_icp'] for ir in irs]
            self.position_results[pos]['dist_position_kp'] = [ir['dist_position']['kp'] for ir in irs if 'kp' in ir['dist_position']]
            self.position_results[pos]['dist_position_kp_icp'] = [ir['dist_position']['kp_icp'] for ir in irs if 'kp' in ir['dist_position']]

            self.position_results[pos]['ADD_nn'] = [ir['ADD_nn'] for ir in irs]
            self.position_results[pos]['ADD_nn_icp'] = [ir['ADD_nn_icp'] for ir in irs]
            self.position_results[pos]['ADD_kp'] = [ir['ADD_kp'] for ir in irs if 'ADD_kp' in ir]
            self.position_results[pos]['ADD_kp_icp'] = [ir['ADD_kp_icp'] for ir in irs if 'ADD_kp' in ir]

            if _config.TEST.SEGMENTATION.evaluate:
                self.position_results[pos]['segmentation_accuracy'] = [ir['segmentation']['accuracy'] for ir in irs]
                self.position_results[pos]['segmentation_precision'] = [ir['segmentation']['precision'] for ir in irs]
                self.position_results[pos]['segmentation_recall'] = [ir['segmentation']['recall'] for ir in irs]

                for cls in _config.INFERENCE.SEGMENTATION.classes:
                    self.position_results[pos][f'segmentation_{cls}_precision'] = [ir['segmentation']['class_results'][cls]['precision'] for ir in irs]
                    self.position_results[pos][f'segmentation_{cls}_recall'] = [ir['segmentation']['class_results'][cls]['recall'] for ir in irs]

        for prs in self.position_results.values():
            for k in prs:
                if len(prs[k]) > 0:
                    self.overall_results[k].append(statistics.mean(prs[k]))

        self.overall_results["calibration_angle_diff"] = -100
        self.overall_results["calibration_dist_position"] = -100
        if self.calibration.pose_camera_link is not None:
            calibration_metrics = metrics.compute_pose_metrics(self.calibration.pose_camera_link, self._gt_base_to_cam_pose)
            self.overall_results["calibration_angle_diff"] = calibration_metrics["angle_diff"]
            self.overall_results["calibration_dist_position"] = calibration_metrics["dist_position"]

        self.export_to_xslx()

    def _put_values_for_col(self, sheet, col_id, start_row, values):
        sheet.cell(row=start_row, column=col_id).value = round(statistics.mean(values), 4) if len(values) > 0 else "N/A"
        sheet.cell(row=start_row + 1, column=col_id).value = round(min(values), 4) if len(values) > 0 else "N/A"
        sheet.cell(row=start_row + 2, column=col_id).value = round(max(values), 4) if len(values) > 0 else "N/A"
        sheet.cell(row=start_row + 3, column=col_id).value = round(statistics.median(values), 4) if len(values) > 0 else "N/A"
        sheet.cell(row=start_row + 4, column=col_id).value = round(statistics.stdev(values), 4) if len(values) > 1 else "N/A"

    def _create_excel_cells(self, sheet, title, results, start_cell="A-1"):
        start_cell = start_cell.split('-')
        col = start_cell[0].upper()
        col_id = ord(col) - ord('A') + 1
        seg_col_id = col_id + 16
        row = int(start_cell[1])

        sheet.merge_cells(f'{chr(ord(col) - 1)}{row}:{chr(ord(col) - 1)}{row + 7}')
        sheet.cell(row=row, column=1).value = title
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", textRotation=90)

        sheet.merge_cells(f'{col}{row}:{col}{row+2}')
        sheet.merge_cells(f'{chr(ord(col) + 1)}{row}:{chr(ord(col) + 4)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + 5)}{row}:{chr(ord(col) + 8)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + 9)}{row}:{chr(ord(col) + 12)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + 13)}{row}:{chr(ord(col) + 13)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + 14)}{row}:{chr(ord(col) + 15)}{row + 1}')

        sheet.merge_cells(f'{chr(ord(col) + seg_col_id - col_id)}{row}:{chr(ord(col) + seg_col_id - col_id + 8)}{row}')
        sheet.merge_cells(f'{chr(ord(col) + seg_col_id - col_id)}{row + 1}:{chr(ord(col) + seg_col_id - col_id + 2)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + seg_col_id - col_id + 3)}{row + 1}:{chr(ord(col) + seg_col_id - col_id + 4)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + seg_col_id - col_id + 5)}{row + 1}:{chr(ord(col) + seg_col_id - col_id + 6)}{row + 1}')
        sheet.merge_cells(f'{chr(ord(col) + seg_col_id - col_id + 7)}{row + 1}:{chr(ord(col) + seg_col_id - col_id + 8)}{row + 1}')

        sheet.merge_cells(f'{chr(ord(col) - 1)}{row + 8}:{chr(ord(col) + seg_col_id - col_id + 8)}{row + 8}')
        sheet.cell(row=row+8, column=1).fill = PatternFill("solid", fgColor="DDDDDD")

        sheet.cell(row=row+3, column=col_id).value = 'Avg'
        sheet.cell(row=row+4, column=col_id).value = 'Min'
        sheet.cell(row=row+5, column=col_id).value = 'Max'
        sheet.cell(row=row+6, column=col_id).value = 'Med'
        sheet.cell(row=row+7, column=col_id).value = 'Std'

        sheet.cell(row=row, column=col_id + 1).value = 'Translation (Euler Dist - m)'
        sheet.cell(row=row+2, column=col_id + 1).value = 'NN'
        self._put_values_for_col(sheet, col_id + 1, row + 3, results['dist_position_nn'])
        sheet.cell(row=row+2, column=col_id + 2).value = 'NN + ICP'
        self._put_values_for_col(sheet, col_id + 2, row + 3, results['dist_position_nn_icp'])
        sheet.cell(row=row + 2, column=col_id + 3).value = 'KP'
        self._put_values_for_col(sheet, col_id + 3, row + 3, results['dist_position_kp'])
        sheet.cell(row=row + 2, column=col_id + 4).value = 'KP + ICP'
        self._put_values_for_col(sheet, col_id + 4, row + 3, results['dist_position_kp_icp'])

        sheet.cell(row=row, column=col_id + 5).value = 'Rotation (Angle Diff - rad)'
        sheet.cell(row=row + 2, column=col_id + 5).value = 'NN'
        self._put_values_for_col(sheet, col_id + 5, row + 3, results['angle_diff_nn'])
        sheet.cell(row=row + 2, column=col_id + 6).value = 'NN + ICP'
        self._put_values_for_col(sheet, col_id + 6, row + 3, results['angle_diff_nn_icp'])
        sheet.cell(row=row+2, column=col_id + 7).value = 'KP'
        self._put_values_for_col(sheet, col_id + 7, row + 3, results['angle_diff_kp'])
        sheet.cell(row=row+2, column=col_id + 8).value = 'KP + ICP'
        self._put_values_for_col(sheet, col_id + 8, row + 3, results['angle_diff_kp_icp'])

        sheet.cell(row=row, column=col_id + 9).value = 'ADD (m)'
        sheet.cell(row=row + 2, column=col_id + 9).value = 'NN'
        self._put_values_for_col(sheet, col_id + 9, row + 3, results['ADD_nn'])
        sheet.cell(row=row + 2, column=col_id + 10).value = 'NN + ICP'
        self._put_values_for_col(sheet, col_id + 10, row + 3, results['ADD_nn_icp'])
        sheet.cell(row=row+2, column=col_id + 11).value = 'KP'
        self._put_values_for_col(sheet, col_id + 11, row + 3, results['ADD_kp'])
        sheet.cell(row=row+2, column=col_id + 12).value = 'KP + ICP'
        self._put_values_for_col(sheet, col_id + 12, row + 3, results['ADD_kp_icp'])

        sheet.cell(row=row, column=col_id + 13).value = 'Key Points'
        sheet.cell(row=row+2, column=col_id + 13).value = 'Distance Error (m)'
        self._put_values_for_col(sheet, col_id + 13, row + 3, results['mean_kp_error'])

        sheet.cell(row=row, column=col_id + 14).value = 'Base2Cam Errors\r\n(NN+ICP)'
        sheet.cell(row=row, column=col_id + 14).alignment = Alignment(wrap_text=True)
        sheet.cell(row=row+2, column=col_id + 14).value = 'Translation (m)'
        self._put_values_for_col(sheet, col_id + 14, row + 3, results['base2cam_dist_position'])
        sheet.cell(row=row+2, column=col_id + 15).value = 'Rotation (rad)'
        self._put_values_for_col(sheet, col_id + 15, row + 3, results['base2cam_angle_diff'])

        if _config.TEST.SEGMENTATION.evaluate:
            sheet.cell(row=row, column=seg_col_id).value = 'Segmentation'
            sheet.cell(row=row+1, column=seg_col_id).value = 'All'
            sheet.cell(row=row+2, column=seg_col_id).value = 'Accuracy'
            self._put_values_for_col(sheet, seg_col_id, row + 3, results['segmentation_accuracy'])
            sheet.cell(row=row+2, column=seg_col_id + 1).value = 'Precision'
            self._put_values_for_col(sheet, seg_col_id + 1, row + 3, results['segmentation_precision'])
            sheet.cell(row=row+2, column=seg_col_id + 2).value = 'Recall'
            self._put_values_for_col(sheet, seg_col_id + 2, row + 3, results['segmentation_recall'])
            sheet.cell(row=row+1, column=seg_col_id + 3).value = 'End Effector'
            sheet.cell(row=row+2, column=seg_col_id + 3).value = 'Precision'
            self._put_values_for_col(sheet, seg_col_id + 3, row + 3, results['segmentation_ee_precision'])
            sheet.cell(row=row+2, column=seg_col_id + 4).value = 'Recall'
            self._put_values_for_col(sheet, seg_col_id + 4, row + 3, results['segmentation_ee_recall'])
            sheet.cell(row=row+1, column=seg_col_id + 5).value = 'Arm'
            sheet.cell(row=row+2, column=seg_col_id + 5).value = 'Precision'
            self._put_values_for_col(sheet, seg_col_id + 5, row + 3, results['segmentation_arm_precision'])
            sheet.cell(row=row+2, column=seg_col_id + 6).value = 'Recall'
            self._put_values_for_col(sheet, seg_col_id + 6, row + 3, results['segmentation_arm_recall'])
            sheet.cell(row=row+1, column=seg_col_id + 7).value = 'Background'
            sheet.cell(row=row+2, column=seg_col_id + 7).value = 'Precision'
            self._put_values_for_col(sheet, seg_col_id + 7, row + 3, results['segmentation_background_precision'])
            sheet.cell(row=row+2, column=seg_col_id + 8).value = 'Recall'
            self._put_values_for_col(sheet, seg_col_id + 8, row + 3, results['segmentation_background_recall'])

        max_col_id = seg_col_id + 8
        max_row_id = row + 7

        return max_row_id, max_col_id

    def export_to_xslx(self):
        wb = openpyxl.Workbook()
        sheet = wb.active

        border = Border(
            left=Side(border_style='thin', color="AAAAAA"),
            right=Side(border_style='thin', color="AAAAAA"),
            top=Side(border_style='thin', color="AAAAAA"),
            bottom=Side(border_style='thin', color="AAAAAA")
        )

        max_row, max_col = self._create_excel_cells(sheet, "OVERALL", self.overall_results, start_cell="B-5")
        for row in range(5, max_row + 1):
            for col in range(1, max_col + 1):
                sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFECB3")
                sheet.cell(row=row, column=col).border = border

        for i, (pk, prs) in enumerate(self.position_results.items()):
            self._create_excel_cells(sheet, pk, prs, start_cell=f"B-{5 + (i + 1) * 9}")

        sheet.merge_cells(f'B1:{chr(ord("A") + max_col - 1)}1')
        sheet.cell(row=1, column=1).value = 'Config: (for reproduction)'
        sheet.cell(row=1, column=2).value = json.dumps(_config())

        # Fill calibration data
        sheet.merge_cells(f'A2:B3')
        sheet.merge_cells(f'C2:D2')
        sheet.merge_cells(f'C3:D3')
        sheet.merge_cells(f'F2:H2')
        sheet.merge_cells(f'F3:H3')
        sheet.merge_cells(f'I2:{chr(ord("A") + max_col - 1)}2')
        sheet.merge_cells(f'I3:{chr(ord("A") + max_col - 1)}3')

        sheet.merge_cells(f'A4:{chr(ord("A") + max_col - 1)}4')
        sheet.cell(row=4, column=1).fill = PatternFill("solid", fgColor="DDDDDD")

        sheet.cell(row=2, column=1).value = 'CALIBRATION'
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

        sheet.cell(row=2, column=3).value = 'Translation Error (m):'
        sheet.cell(row=3, column=3).value = 'Rotation Error (rad):'
        sheet.cell(row=2, column=5).value = round(self.overall_results["calibration_dist_position"] or 0, 4) or "N/A"
        sheet.cell(row=3, column=5).value = round(self.overall_results["calibration_angle_diff"] or 0, 4) or "N/A"

        sheet.cell(row=2, column=6).value = 'GT Pose (x, y, z, qw, qx, qy, qz):'
        sheet.cell(row=3, column=6).value = 'PRED Pose (x, y, z, qw, qx, qy, qz):'
        sheet.cell(row=2, column=9).value = ', '.join([f'{a:.4f}' for a in  self._gt_base_to_cam_pose])
        sheet.cell(row=3, column=9).value = 'N/A'
        if self.calibration.pose_camera_link is not None:
            sheet.cell(row=3, column=9).value = ', '.join([f'{a:.4f}' for a in  self.calibration.pose_camera_link])

        wb.save(_config.TEST.output)


if __name__ == "__main__":
    random.seed(_config.TEST.seed)
    np.random.seed(_config.TEST.seed)
    torch.manual_seed(_config.TEST.seed)

    app = TestApp()
    app.run_tests()
    # app.export_to_xslx()

    # ipdb.set_trace()
